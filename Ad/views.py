from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponseRedirect, JsonResponse
from login.models import *
from Ad.models import *
from client.models import *
from Qc.models import *
from django.shortcuts import redirect
from django.urls import reverse
import openpyxl 
from openpyxl import load_workbook


# Create your views here.

def dashboard(request):
    return render (request,'dashboard.html')

def dashboardtabel(request):
    return render (request,'dashboardtabel.html')

def admin(request):
    con = contract.objects.all()
    return render (request,'admin.html',{'con':con})

def add_contractor(request):
    if request.method =='POST':
        print(request.POST.copy())
        contractor =request.POST.get('contractor')
        contractortype =request.POST.get('contractor_role')
        
        contractorprocess = contract(contractor=contractor,contractortype=contractortype)
        contractorprocess.save()
        return redirect ('add_contractor')
    else:
        return render (request,'add_contractor.html')

def projectlist(request):
    if request.method == "POST":
        response = redirect('dashboard')  # Replace 'some_view_name' with the name of the view you want to redirect to

        # Set a cookie named 'my_cookie' with the value 'my_value'
        response.set_cookie('active_project',request.POST.get('project_id'), max_age=36000000)  # The cookie will last for 1 hour

        return response
    user = request.user
    if user.role=="Admin":


        Adpro = ad.objects.get(user=user)
        print(Adpro)
        pro =Adpro.pro.all()
    elif user.role=="Qc":
        qcpro =Qc.objects.get(user=user)
        print(qcpro)
        pro =qcpro.pro.all()
    elif user.role=="Client":
        clientpro =client.objects.get(user=user)
        print(clientpro)
        pro =clientpro.pro.all()

    

    return render (request,'projectlist.html',{'data':pro})

def startproject(request):
     return render (request,'startproject.html')

def createproject(request):
    data = project.objects.all()
    return render (request,'createproject.html', {'data': data})

def newproject(request):
    data = project.objects.all()
    main = contract.objects.filter(contractortype = "Main contractor")
    sub = contract.objects.filter(contractortype = "sub contractor")
    if request.method =='POST':
        print(request.POST.copy())
        print(request.FILES)
        projectnum   =  request.POST.get('projectnum')
        projectname =request.POST.get('projectname')
        customer    =request.POST.get('customer')
        role    = request.POST.get('role')
        maincontra =request.POST.get('main')
        subcontractor=request.POST.get('subcontractor')
        client     =request.POST.get('client')
        startdate  =request.POST.get('sdate')
        enddate    =request.POST.get('edate')
        fabcont    =request.POST.get('fab')
        fabcli    =request.POST.get('fabclient')
        finalacc  =request.POST.get('final')
        finalacp  =request.POST.get('finalacp')
        finalacct =request.POST.get('finalacct')
        drawrev   =request.POST.get('draw')
        weldsum  =request.POST.get('weldsum')
        dailyrec  =request.POST.get('dailyrec')
        indrec   =request.POST.get('indrec')
        spoolre  =request.POST.get('spool')
        ndereq   =request.POST.get('ndeeq')
        rtform   =request.POST.get('rtform')
        ptform   =request.POST.get('ptform')
        pmiform  =request.POST.get('pmiform')
        pwhtfprm =request.POST.get('pwhtform')
        mtform   =request.POST.get('mtform')
        bhnform  =request.POST.get('bhnform')
        ferrite  =request.POST.get('ferform')
        pautform =request.POST.get('pautform')
        prhtform =request.POST.get('prhtform')
        utform   =request.POST.get('utform') 
        utgform  =request.POST.get('utgform') 
        shopadd  =request.POST.get('shopadd')
        fieldadd =request.POST.get('fieldadd')
        cliconnum=request.POST.get('clientcon')
        utocjobno =request.POST.get('utocjob')
        weld_id = request.POST.get('weld_id')
        projectclosed=request.POST.get('projectclosed')
        print(projectnum,"utoprojectnumcjobno")

        pro = project(projectno=projectnum,projectname=projectname,customer=customer,subcontractor= subcontractor,
                      role=role,maincont=maincontra,client=client,startdate=startdate,enddate=enddate,fabcont=fabcont,
                      fabcli=fabcli, owner_logo=request.FILES.get("owner_logo"), client_logo=request.FILES.get("client_logo"), utoc_logo=request.FILES.get("utoc_logo"),
                       finalaccon=finalacc,finalaccus=finalacp,finalac=finalacct,drarev=drawrev,weldsum=weldsum,dailyrec=dailyrec,
                        indrec=indrec,spoolre=spoolre,rtform=rtform,ndereq=ndereq,ptform=ptform,pmiform=pmiform,pwhtform=pwhtfprm,
                         mtform=mtform,bhnform=bhnform,ferrite=ferrite,pautform=pautform,prhtform=prhtform,utform=utform,utgform=utgform,
                          shopadd=shopadd,fieldadd=fieldadd,cliconno=cliconnum,utocjobno=utocjobno,weldid=weld_id,Projectclosed=projectclosed)
        print(projectnum,"projectnum")
        pro.save()
        return render (request,'createproject.html', {'data': data} )
    else:
        return render (request,'newproject.html',{'main':main,'sub':sub})   



    

def detailproject(request, pro_id):
    ad_Detail = project.objects.get(id=pro_id)
    return render (request,'detailproject.html',{'data':ad_Detail})

def project_welder(request,id):
    projectId = id
    pro= project.objects.get(id=id)
    prowelders =pro.welders.all()
    return render (request,'project_welder.html',{'proId':projectId,'prowelder':prowelders})

def addproject_welder(request,id):
    con = contract.objects.filter(contractortype = "sub contractor")
    if request.method == "GET":

        dropdown_value1 = request.GET.get('dropdown1', '')

        welde = welder.objects.all()

        if dropdown_value1:
            welde = welder.objects.filter(contrac = dropdown_value1)
    else:
        pro = project.objects.get(id=id)
        print(pro)
        check1 = request.POST.getlist('prowelder')
        print(check1)
        for area_code in check1:
                proj = welder.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                pro.welders.add(proj)
        return redirect(reverse('project_welder', args=(id,)))

    return render (request,'addproject_welder.html', {'proId':id,'con':con,'welders':welde})

def setupoption(request):
    return render (request,'setupoption.html')

def addweldoption(request):
    return render (request,'addweldoption.html')

def weldmapoption(request):
    return render (request,'weldmapoption.html')

def userprofile(request):
    data1 = ad.objects.all()
    data2 = Qc.objects.all()
    data3 = client.objects.all()
    data4 = Fitter.objects.all()
    pro = project.objects.all()
    pro1=ad.objects.all()
    data_dict = {
        'admin_data': data1,
        'qc_data': data2,
        'client_data': data3,
        'fitter_data': data4,
        'project': pro,
        'project1':pro1
    }
    return render (request,'userprofile.html',{'data_dict':data_dict})

def newprofile(request):
    data1 = ad.objects.all()
    data2 = Qc.objects.all()
    data3 = client.objects.all()
    pro = project.objects.all()
    pro1=ad.objects.all()
    data_dict = {
        'admin_data': data1,
        'qc_data': data2,
        'client_data': data3
    }

    if request.method == 'POST':

        role = request.POST.get('role')

        print(request.POST.copy())

        emp_name = request.POST.get('Name')
        userid=request.POST.get('userid')
        emp_email = request.POST.get('mail')
        gender = request.POST.get('gender')
        date = request.POST.get('date')
        check1 = request.POST.getlist('project')
        empNum = request.POST.get('num')
        role = request.POST.get('role')
        password = "123"
        print(check1)

        user = User()
        user.first_name = emp_name
        user.username = emp_name
        user.email = emp_email
        user.role = role
        user.set_password(password)
        user.save()
        link_id = user.id
            
        if role =="Admin":
            new_ad = ad( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender, user_id =link_id,userid=userid) # model_Name(field_Name = Variable_Name)
            new_ad.save()
            for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                new_ad.pro.add(proj)

        elif role == "Qc":
         new_qc = Qc( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender, user_id = link_id, userid=userid) # model_Name(field_Name = Variable_Name)
         new_qc.save()
         for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                new_qc.pro.add(proj)
        elif role == "Client":
         new_client = client(name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender, user_id = link_id)
         new_client.save()
         for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                new_client.pro.add(proj)
        elif role == "Fitter":
         new_fitter = Fitter( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender, user_id = link_id, userid=userid) # model_Name(field_Name = Variable_Name)
         new_fitter.save()
         for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                new_fitter.pro.add(proj)
         messages.success(request,"Fitter Added..!!!")
        return redirect ('userprofile')

        # return render (request,'userprofile.html',{'data_dict':data_dict})
    else:
        return render (request,'newprofile.html',{'pro':pro})
    
def newprofileadedit(request, id):
    data1 = ad.objects.get(id=id)
    print(data1)
    pro = project.objects.all()
    tagged_projects = data1.pro.all()
    
    if request.method == 'POST':
        print(request.POST.copy())

        emp_name = request.POST.get('Name')
        userid=request.POST.get('userid')
        emp_email = request.POST.get('mail')
        gender = request.POST.get('gender')
        date = request.POST.get('date')
        check1 = request.POST.getlist('project')
        empNum = request.POST.get('num')

        data1.user.first_name = emp_name
        data1.user.username = emp_name
        data1.user.email = emp_email
        data1.user.save()
        ad.objects.filter(id=id).update( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender,userid=userid) # model_Name(field_Name = Variable_Name)
        if check1:
            # Clear existing areacode associations
            data1.pro.clear()
            for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                data1.pro.add(proj)
        messages.success(request,"Admin Edited..!!!")
  
        return redirect ('userprofile')
    else:
        return render (request,'newprofile.html', {'data':data1, 'pro':pro, 'tagged_projects': tagged_projects})
    
def newprofileaddel(request, id):
    


        addel = ad.objects.get(id = id)
        print(addel)
        addel.delete()
        return redirect('userprofile')


def newprofileqcedit(request, id):
    data1 = Qc.objects.get(id=id)
    pro=project.objects.all()
    tagged_projects=data1.pro.all()
    if request.method == 'POST':
        print(request.POST.copy())

        emp_name = request.POST.get('Name')
        userid=request.POST.get('userid')
        emp_email = request.POST.get('mail')
        gender = request.POST.get('gender')
        date = request.POST.get('date')
        check1 = request.POST.getlist('project')
        empNum = request.POST.get('num')
        

        data1.user.first_name = emp_name
        data1.user.username = emp_name
        data1.user.email = emp_email
        data1.user.save()
        Qc.objects.filter(id=id).update( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender,userid=userid) # model_Name(field_Name = Variable_Name)
        if check1:
            # Clear existing areacode associations
            data1.pro.clear()
            for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                data1.pro.add(proj)
        messages.success(request,"Qc Edited..!!!")
  
        return redirect ('userprofile')
    else:
        return render (request,'newprofile.html', {'data':data1,"pro":pro,"tagged_projects":tagged_projects})
    
def newprofileqcdel(request,id):
        qcdel=Qc.objects.get(id=id)
        print(qcdel)
        qcdel.delete()
        return redirect('userprofile')

def newprofilefitteredit(request, id):
    data1 = Fitter.objects.get(id=id)
    pro=project.objects.all()
    tagged_projects=data1.pro.all()
    if request.method == 'POST':
        print(request.POST.copy())

        emp_name = request.POST.get('Name')
        userid=request.POST.get('userid')
        emp_email = request.POST.get('mail')
        gender = request.POST.get('gender')
        date = request.POST.get('date')
        check1 = request.POST.getlist('project')
        empNum = request.POST.get('num')
        

        data1.user.first_name = emp_name
        data1.user.username = emp_name
        data1.user.email = emp_email
        data1.user.save()
        Fitter.objects.filter(id=id).update( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender,userid=userid) # model_Name(field_Name = Variable_Name)
        if check1:
            # Clear existing areacode associations
            data1.pro.clear()
            for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                data1.pro.add(proj)
        messages.success(request,"Qc Edited..!!!")
  
        return redirect ('userprofile')
    else:
        return render (request,'newprofile.html', {'data':data1,"pro":pro,"tagged_projects":tagged_projects})
    
def newprofilefitterdel(request,id):
        fitterdel=Fitter.objects.get(id=id)
        print(fitterdel)
        fitterdel.delete()
        return redirect('userprofile')

def detailprofilefitter(request,id):
    detail=Fitter.objects.get(id=id)
    tagged_projects = detail.pro.all()
    print(detail)

    return render (request,'detailprofile.html',{'data':detail, 'tagged_projects': tagged_projects })

def newprofileclientedit(request, id):
    data1 = client.objects.get(id=id)
    pro=project.objects.all()
    tagged_projects=data1.pro.all()
    if request.method == 'POST':
        
        print(request.POST.copy())

        emp_name = request.POST.get('Name')
        userid=request.POST.get('userid')
        emp_email = request.POST.get('mail')
        gender = request.POST.get('gender')
        date = request.POST.get('date')
        check1 = request.POST.getlist('project')
        empNum = request.POST.get('num')

        data1.user.first_name = emp_name
        data1.user.username = emp_name
        data1.user.email = emp_email
        data1.user.save()
        client.objects.filter(id=id).update( name = emp_name, number = empNum, DOB =date, email =emp_email,gender = gender,userid=userid) # model_Name(field_Name = Variable_Name)
        if check1:
            # Clear existing areacode associations
            data1.pro.clear()
            for area_code in check1:
                proj = project.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                data1.pro.add(proj)
        messages.success(request,"Client Edited..!!!")
  
        return redirect ('userprofile')
    else:
        return render (request,'newprofile.html', {'data':data1,"pro":pro,"tagged_projects":tagged_projects}) 
     
def newprofileclientdel(request,id):
        clientdel=client.objects.get(id=id)
        print(clientdel)
        clientdel.delete()
        return redirect('userprofile')     

def detailprofileqc(request,proid):
    detail=Qc.objects.get(id=proid)
    tagged_projects = detail.pro.all()
    print(detail)

    return render (request,'detailprofile.html',{'data':detail, 'tagged_projects': tagged_projects })

def detailprofilead(request,proid):
    detail=ad.objects.get(id=proid)
    tagged_projects = detail.pro.all()
    print(detail)

    return render (request,'detailprofile.html',{'data':detail,'tagged_projects': tagged_projects})

def detailprofileclient(request,proid):
    detail=client.objects.get(id=proid)
    tagged_projects = detail.pro.all()
    print(detail)

    return render (request,'detailprofile.html',{'data':detail, 'tagged_projects': tagged_projects})

def project_edit(request, id):
    data1 = project.objects.get(id=id)
    main = contract.objects.filter(contractortype = "Main contractor")
    sub = contract.objects.filter(contractortype = "sub contractor")
    print(data1)
    if request.method == 'POST':
        print(request.POST.copy())
        projectnum   =  request.POST.get('projectnum')
        projectname =request.POST.get('projectname')
        customer    =request.POST.get('customer')
        role    = request.POST.get('role')
        maincontra =request.POST.get('main')
        subcontractor=request.POST.get('subcontractor')
        client     =request.POST.get('client')
        startdate  =request.POST.get('sdate')
        enddate    =request.POST.get('edate')
        fabcont    =request.POST.get('fab')
        fabcli    =request.POST.get('fabclient')
        finalacc  =request.POST.get('final')
        finalacp  =request.POST.get('finalacp')
        finalacct =request.POST.get('finalacct')
        drawrev   =request.POST.get('draw')
        weldsum  =request.POST.get('weldsum')
        dailyrec  =request.POST.get('dailyrec')
        indrec   =request.POST.get('indrec')
        spoolre  =request.POST.get('spool')
        ndereq   =request.POST.get('ndeeq')
        rtform   =request.POST.get('rtform')
        ptform   =request.POST.get('ptform')
        pmiform  =request.POST.get('pmiform')
        pwhtfprm =request.POST.get('pwhtform')
        mtform   =request.POST.get('mtform')
        bhnform  =request.POST.get('bhnform')
        ferrite  =request.POST.get('ferform')
        pautform =request.POST.get('pautform')
        prhtform =request.POST.get('prhtform')
        utform   =request.POST.get('utform') 
        utgform  =request.POST.get('utgform') 
        shopadd  =request.POST.get('shopadd')
        fieldadd =request.POST.get('fieldadd')
        cliconnum=request.POST.get('clientcon')
        utocjobno =request.POST.get('utocjob')
        weld_id = request.POST.get('weld_id')
        projectclosed=request.POST.get('projectclosed')
        
        project.objects.filter(id=id).update(projectno=projectnum,projectname=projectname,customer=customer,subcontractor=subcontractor,
                        role=role,maincont=maincontra,client=client,startdate=startdate,enddate=enddate,fabcont=fabcont,
                        fabcli=fabcli,finalaccon=finalacc,finalaccus=finalacp,finalac=finalacct,drarev=drawrev,weldsum=weldsum,dailyrec=dailyrec,
                            indrec=indrec,spoolre=spoolre,rtform=rtform,ndereq=ndereq,ptform=ptform,pmiform=pmiform,pwhtform=pwhtfprm,
                            mtform=mtform,bhnform=bhnform,ferrite=ferrite,pautform=pautform,prhtform=prhtform,utform=utform,utgform=utgform,
                            shopadd=shopadd,fieldadd=fieldadd,cliconno=cliconnum,utocjobno=utocjobno,weldid=weld_id,Projectclosed=projectclosed,
                            owner_logo=request.FILES.get("owner_logo"), client_logo=request.FILES.get("client_logo"), utoc_logo=request.FILES.get("utoc_logo"))

        project_instance = project.objects.filter(id=id).first()
        # Update image fields if new files are provided
        if request.FILES.get("owner_logo"):
            project_instance.owner_logo = request.FILES.get("owner_logo")
        if request.FILES.get("client_logo"):
            project_instance.client_logo = request.FILES.get("client_logo")
        if request.FILES.get("utoc_logo"):
            project_instance.utoc_logo = request.FILES.get("utoc_logo")

        # Save the instance
        project_instance.save()
        messages.success(request, "Project Edited Succesfully")

        return redirect('createproject')
    else:
        return render (request,'newproject.html', {'data1':data1,'main':main,'sub':sub})
def projectdel(request,id):
    delproject=project.objects.get(id=id)
    delproject.delete()
    return redirect("createproject")


def weldtypes(request):
     data = Weld.objects.all()
     return render (request,'weldtypes.html',{'data':data})

def newweldtype(request):
    data = Weld.objects.all()
    if request.method == 'POST':
        print(request.POST.copy())
        weld_type = request.POST.get('weldtype') 
        weld_desc = request.POST.get('typedesc')

        new_weld = Weld (weldType = weld_type, weldDesc = weld_desc)
        print(weld_type,"matrixx")
        new_weld.save()
        return redirect( 'weldtypes' )
    else:
        return render (request,'newweldtype.html')
    
def weld_edit(request, id):
    data = Weld.objects.get(id=id)
    print(data)
    if request.method == 'POST':
        print(request.POST.copy())
        weld_type = request.POST.get('weldtype') 
        weld_desc = request.POST.get('typedesc')

        Weld.objects.filter(id=id).update(weldType=weld_type, weldDesc = weld_desc)
        messages.success(request,"Weld Edited Succesfully")

        return redirect('weldtypes')
    else:
        return render (request,'newweldtype.html', {'data':data} )

    
def delete_weld(request, id):
    weld_delete = Weld.objects.get(id = id)  
    weld_delete.delete()
    return redirect('weldtypes')  

def weldprocess(request):
    data= WeldProcess.objects.all()
    return render (request,'weldprocess.html',{'data': data} )

def newweldprocess(request):
    if request.method =='POST':
        print(request.POST.copy())
        weld_process  = request.POST.get('weldprocess')
        weld_process_description =request.POST.get('welddescription')


        new_weld_process = WeldProcess (weldProcess=weld_process, weldProcessDesc= weld_process_description)
        new_weld_process.save()
        return redirect ('weldprocess')
    else:
         return render (request,'newweldprocess.html')
    
def weld_process_edit(request,id):
    data=WeldProcess.objects.get(id=id)
    print(data)
    if request.method =='POST':
        print(request.POST.copy())
        weld_process  = request.POST.get('weldprocess')
        weld_process_description =request.POST.get('welddescription')
        WeldProcess.objects.filter(id=id).update(weldProcess=weld_process, weldProcessDesc= weld_process_description)
        messages.success(request,"weldprocess edit sucessfully")

        return redirect("weldprocess")
    else:
        return render (request,'newweldprocess.html',{'data':data})
    
def delete_weld_process(request, id):
    weld_process_delete = WeldProcess.objects.get(id = id)  
    weld_process_delete.delete()
    return redirect('weldprocess')  





def weldlocation(request):
    data=weld_location.objects.all()
    return render (request,'weldlocation.html',{"data":data})

def newweldlocation(request):

    if request.method =='POST':
          print(request.POST.copy())
          weldinglocation =request.POST.get('weldlocation')
          weldinglocationtype =request.POST.get('WeldDescrption')
          role=request.POST.get('choice')
          newweldlocationprocess =  weld_location(weldlocation =weldinglocation,weldlocationdesc =weldinglocationtype, role = role)
          newweldlocationprocess.save()
          return redirect ('weldlocation')
    else:
         return render (request,'newweldlocation.html')
    
def weld_location_edit(request,id):
    data=weld_location.objects.get(id=id)
    print(data)
    if request.method =='POST':
        print(request.POST.copy())
        weldlocation =request.POST.get('weldlocation')
        weldlocationtype =request.POST.get('WeldDescrption')
        role=request.POST.get('choice')
        
        weld_location.objects.filter(id=id).update(weldlocation=weldlocation,weldlocationdesc=weldlocationtype,role=role)
        messages.success(request,"weldlocation edit sucessfully")

        return redirect("weldlocation")
    else:
        return render (request,'newweldlocation.html',{'data':data})
    
def delete_weld_location(request, id):
    weld_location_delete = weld_location.objects.get(id = id)  
    weld_location_delete.delete()
    return redirect('weldlocation')  

def welders(request):
    data = welder.objects.all()
    return render (request,'welders.html',{'data':data})

def new_welder_wps(request):
    wps = w_p_s.objects.all()
    return render('newwelders', {'data': wps})

def detailwelders(request,id):
    weldDetail = welder.objects.get(id=id)
    tagged_projects = weldDetail.weldwps.all()
    return render (request,'detailwelders.html',{'data':weldDetail,'tagged_projects':tagged_projects})

def newwelders(request):
    wps = w_p_s.objects.all()
    con = contract.objects.filter(contractortype='sub contractor')
    if request.method =='POST':
        print(request.POST.copy())
        weldername= request.POST.get('welder_name')
        welderid= request.POST.get('welder_id')
        finno= request.POST.get('fin_no')
        primosno= request.POST.get('primos_no')
        conId=request.POST.get('con')
        primosexpdate= request.POST.get('expiry_date')
        check1 = request.POST.getlist('wpsSelect')

        newweldersprocess =welder(weldname=weldername,finno=finno,primosno=primosno,primosexpdate=primosexpdate,welderid=welderid,contrac_id=conId)
        newweldersprocess.save()
        for area_code in check1:
                proj = w_p_s.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                newweldersprocess.weldwps.add(proj)
        return redirect ('welders')
    else:
        return render (request,'newwelders.html',{'wps':wps,'con':con})
def weldersedit(request,id):
    data=welder.objects.get(id=id)
    tagged_projects = data.weldwps.all()
    wps = w_p_s.objects.all()
    print(data)
    if request.method =='POST':
        print(request.POST.copy())
        weldername= request.POST.get('welder_name')
        finno= request.POST.get('fin_no')
        welderid= request.POST.get('welder_id')
        primosno= request.POST.get('primos_no')
        primosexpdate= request.POST.get('expiry_date')
        check1 = request.POST.getlist('wpsSelect')
        

       
        welder.objects.filter(id=id).update(weldname=weldername,finno=finno,primosno=primosno,primosexpdate=primosexpdate,welderid=welderid)
        if check1:
            data.weldwps.clear()
        for area_code in check1:
            if area_code:
                proj = w_p_s.objects.get(id=area_code)
                print(proj,'aaaaaaaaaaaaaaa')
                data.weldwps.add(proj)
        messages.success(request,"welders edit sucessfully")
        return redirect ('welders')
    else:
        return render (request,'newwelders.html',{'data':data,'selected_wps':tagged_projects,'wps':wps})
    
def weldersdel(request,id):
    weldersdel=welder.objects.get(id=id)
    weldersdel.delete()
    return redirect("welders")
    

def materials(request):
     data = mat.objects.all()
     return render (request,'materials.html',{'data':data})

def newmaterials(request):

    if request.method =='POST':
        print(request.POST.copy())
        material =request.POST.get('material')
        material_desc =request.POST.get('materialdescription')
        materialdescription=request.POST.get('materialsdescription2')
        print(materialdescription)
        print(material)
        
        new_material_process = mat(materials= material, materialsdesc = material_desc,materialsdescription=materialdescription)
        new_material_process.save()
        return redirect ('materials')
    else:
        return render (request,'newmaterials.html')
    
def detailmaterials(request,id):
    ad_materials=mat.objects.get(id=id)
    matGrade = materialsgrade.objects.filter(mat_id = id)
    print(ad_materials)
    return render (request,'detailmaterials.html',{'data':ad_materials, 'data1': matGrade})



def materials_grade(request, id):
    if request.method =='POST':
        print(request.POST.copy())
        materialgrade=request.POST.get("materialgrade")
        materialdesc=request.POST.get("gradedesc")
        materialgradeprocess=materialsgrade(materialgrade=materialgrade,materialdesc=materialdesc, mat_id = id)
        materialgradeprocess.save()
        return redirect('detailmaterials',id = id)
    else:
        return render (request,'materialsgrade.html')
    
def detailmaterialgrade(request,id):
    grade=materialsgrade.objects.get(id=id)
    print(grade)
    return render (request,'detailmaterials.html',{'data1':grade})

# def materialed(request,id):
#     data=mat.objects.get(id=id)
#     print(data)
#     if request.method =='POST':
#         print(request.POST.copy())
#         material =request.POST.get('material')
#         material_desc =request.POST.get('materialdescription')
#         materialdescription=request.POST.get('materialsdescription2')
#         print(material)
        
#         mat.objects.filter(id=id).update(materials= material, materialsdesc = material_desc,materialsdescription=materialdescription)
#         messages.success(request,"materials edit sucessfully")
#         return redirect ('materials')
#     else:
#         return render (request,'newmaterials.html',{'data':data})
# def materialdel(request,id):
#     materialsdel=mat.objects.get(id=id)
#     materialsdel.delete()
#     return redirect("materials")



    
def materialsedit(request,id):
    data=mat.objects.get(id=id)
    print(data)
    if request.method =='POST':
        print(request.POST.copy())
        material =request.POST.get('material')
        material_desc =request.POST.get('materialdescription')
        materialdescription=request.POST.get('materialsdescription2')
        print(material)
        
        mat.objects.filter(id=id).update(materials= material, materialsdesc = material_desc,materialsdescription=materialdescription)
        messages.success(request,"materials edit sucessfully")
        return redirect ('materials')
    else:
        return render (request,'newmaterials.html',{'data':data})
def materialsdel(request,id):
    materialsdel=mat.objects.get(id=id)
    materialsdel.delete()
    return redirect("materials")

def materialsgradeedit(request,id):
    data=materialsgrade.objects.get(id=id)
    if request.method=='POST':
        print(request.POST.copy())
        materialgrade=request.POST.get("materialgrade")
        materialdesc=request.POST.get("gradedesc")
        materialsgrade.objects.filter(id=id).update(materialgrade=materialgrade,materialdesc=materialdesc)
        messages.success(request,"materials grade edit successfully")
        return redirect('detailmaterials',data.mat_id)
    else:
        return render (request,'materialsgrade.html',{"data":data})
    
def materialgradesdel(request,id):
    materialsgradedel=materialsgrade.objects.get(id=id)
    materialsgradedel.delete()
    return redirect('detailmaterials',materialsgradedel.mat_id)


def uploadmaterials(request):
    return render (request,'uploadmaterials.html')


def wps(request):
    data = w_p_s.objects.all()
    return render (request,'wps.html',{'data':data})

def addwps(request):
    if request.method =='POST':
        print(request.POST.copy())
        pipesize1 =request.POST.get('size_start')
        pipesize2 =request.POST.get('size_end')
        pipethickness1 =request.POST.get('thickness_start')
        pipethickness2 =request.POST.get('thickness_end')
        material =request.POST.get('materials')
        grades    =request.POST.get('grades')
        wpsno    =request.POST.get('wps_no')
        classid    =request.POST.get('class_id')
        description   =request.POST.get('wps_desc')
        welding_process =request.POST.get('weldProcess')
        

        addwpsprocess=w_p_s (pipesize1=pipesize1,pipesize2=pipesize2,pipethickness1=pipethickness1,pipethickness2=pipethickness2,
                             material_id=material,grade_id=grades,wpsno=wpsno,classid=classid,description=description,weldingprocess=welding_process)
        addwpsprocess.save()
        return redirect ('wps')
    else:
        material = mat.objects.values("id", "materials")
        
        welding_process = WeldProcess.objects.values("id","weldProcess")
        return render (request,'addwps.html', {'data1': material, 'selected_value1': material,
                                               'data3': welding_process, 'selected_value': welding_process })
    
def wpsedit(request,id):
    data = w_p_s.objects.get(id=id)
    if request.method =='POST':
        print(request.POST.copy())
        pipesize1 =request.POST.get('size_start')
        pipesize2 =request.POST.get('size_end')
        pipethickness1 =request.POST.get('thickness_start')
        pipethickness2 =request.POST.get('thickness_end')
        material =request.POST.get('materials')
        grades    =request.POST.get('grades')
        wpsno    =request.POST.get('wps_no')
        classid    =request.POST.get('class_id')
        description   =request.POST.get('wps_desc')
        welding_process =request.POST.get('weldProcess')
        

       

        w_p_s.objects.filter(id=id).update(pipesize1=pipesize1,pipesize2=pipesize2,pipethickness1=pipethickness1,pipethickness2=pipethickness2,
                             material_id=material,grade_id=grades,wpsno=wpsno,classid=classid,description=description,weldingprocess=welding_process)
        messages.success(request,"wps edit sucessfully")
        return redirect ('wps')
    else:
        material = mat.objects.values("id", "materials")
        
        welding_process = WeldProcess.objects.values("id","weldProcess")
        return render (request,'addwps.html', {'data':data,'data1': material, 'selected_value1': material,
                                               'data3': welding_process, 'selected_value': welding_process })

def wpsdel(request,id):
    wpsdelete=w_p_s.objects.get(id=id)
    wpsdelete.delete()
    return redirect("wps")  


def detailwps(request,id):
    wps = w_p_s.objects.get(id=id)
    print(wps)
    return render (request,'detailwps.html', {'data': wps})




def fetch_weldGrade(request):

    material = request.GET.get('material')
    print(material, "scad barcode ----------------")
    
    try:
        product = materialsgrade.objects.filter(mat_id=material)

        product_data = []

        for product_instance in product:
            product_data.append({
                'id': product_instance.id,
                'GradeNmae': product_instance.materialgrade,
                # Include other product details as needed
            })
            print(product_data)
        return JsonResponse({'products': product_data})

    
    except materialsgrade.DoesNotExist:
        data = {'error': 'not found'}
        return JsonResponse(data, status=404)


def qr_scanner(request):
    return render (request,'qr_scanner.html')

def qr_fitup(request):
    return render (request,'qr_fitup.html')

def qr_visual(request):
    return render (request,'qr_visual.html')


def tfa_upload(request,id):
    if request.method == 'POST':
        excel_file = request.FILES.get('upload_tfa')

        if excel_file.name.endswith('.xlsx'):
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Assuming columns are in the order: name, email, phone_number
                permanent_joint_id,plant_name,area,process_unit,location,flange_description,drawing_name_A,sheet_no,fluid,pipe_specification,flange_size,standard,critical_service,reference_no,flange_rating,flange_material,gasket_material,bolt_material,nut_material,no_of_bolts,bolt_diameter,lubricant,bolt_length,tightening_method,required_torque,contracting_company,sub_contracting_company,test_pack_no,test_requirement,test_completion_date,reinstatement_completion_date, project_name, client_name, work_order_no, system_number, temp_joint_id = row
                
                tfa = TFAUpload()
                tfa.permanent_joint_id = permanent_joint_id
                tfa.plant_name = plant_name
                tfa.area = area
                tfa.process_unit = process_unit
                tfa.location = location
                tfa.flange_description = flange_description
                tfa.drawing_name_A = drawing_name_A
                tfa.sheet_no = sheet_no
                tfa.fluid = fluid
                tfa.pipe_specification = pipe_specification
                tfa.flange_size = flange_size
                tfa.standard = standard
                tfa.critical_service = critical_service
                tfa.reference_no = reference_no
                tfa.flange_rating = flange_rating
                tfa.flange_material = flange_material
                tfa.gasket_material = gasket_material
                tfa.bolt_material = bolt_material
                tfa.nut_material = nut_material
                tfa.no_of_bolts = no_of_bolts
                tfa.bolt_diameter = bolt_diameter
                tfa.lubricant = lubricant
                tfa.bolt_length = bolt_length
                tfa.tightening_method = tightening_method
                # tfa.qc_scope = qc_scope
                tfa.required_torque = required_torque
                tfa.contracting_company = contracting_company
                tfa.sub_contracting_company = sub_contracting_company
                tfa.test_pack_no = test_pack_no
                tfa.test_requirement = test_requirement
                tfa.test_completion_date = test_completion_date
                tfa.reinstatement_completion_date = reinstatement_completion_date
                tfa.project_name = project_name
                tfa.client_name = client_name
                tfa.temp_joint_id = temp_joint_id
                tfa.work_order_no = work_order_no
                # tfa.stud_and_nut_coating = stud_and_nut_coating
                tfa.system_number = system_number
                # tfa.vessel_id = vessel_id
                tfa.project = project.objects.get(id=id)
                tfa.save()   
                operator = "-"
                if request.user and request.user.is_authenticated:
                    operator = request.user.first_name
                    
                TFAHistory.objects.create(tfa_upload=tfa,status="untouched", operator=operator)

            messages.success(request, 'Data uploaded successfully.')
            return render(request, 'createproject.html')
        
        else:
            
            messages.error(request, 'Please upload a valid Excel file.')
            return render(request, 'createproject.html')
    else:
        return render (request,'tfa_upload.html')


def upload_spoolgen(request,id):
    if request.method == 'POST':
            excel_file = request.FILES.get('Spoolgen_data')

            if excel_file.name.endswith('.xlsx'):
                # Load the Excel workbook
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Assuming columns are in the order: name, email, phone_number
                    PWHT_REQMT,RT_REQMT,MATERIAL,PIPING_CLASS,SERVICE,ACCEPTANCE_STD,LINE_NO,ISO_DRG_NO,Rev,Sheet_No,Spool_No,Weld_No,Type,LOCATION,Size_DB,THK,BATCH,Unit,AREA,Insulation,Paint_Sys,P_ID_no,T_Press_BARG,T_Media_W_G_V_S,MOM_Req,BHN_REQMT,FERRITE_REQMT,MT_PT_REQMT,PMI_REQMT,TIE_IN_NO,OPTTEMP,SUB_PROJECT_NO,PC_No= row
                    if not ISO_DRG_NO:
                        continue
                    if mat.objects.filter(materials=MATERIAL).exists():
                        mat_id = mat.objects.filter(materials=MATERIAL).first()
                    else:
                        material_obj = mat()
                        material_obj.materials = MATERIAL
                        material_obj.save()
                        mat_id = material_obj.id
                    # Save data to sims model
                    spoolgenmaterialsuploadss.objects.create(pwht=PWHT_REQMT,rt=RT_REQMT,material=MATERIAL,piping=PIPING_CLASS,
                                                             service=SERVICE,acceptance=ACCEPTANCE_STD,lineno=LINE_NO,
                                                             drgno=ISO_DRG_NO,rev=Rev,sheetno=Sheet_No,spoolno=Spool_No,weldno=Weld_No,
                                                             type=Type,location=LOCATION,size=Size_DB,thk=THK,batch=BATCH,unit=Unit,area=AREA,insulation=Insulation,paint=Paint_Sys,
                                                             pidno=P_ID_no,tpress=T_Press_BARG,tmedia=T_Media_W_G_V_S,mom=MOM_Req,bhn=BHN_REQMT,ferraite=FERRITE_REQMT,mtpt=MT_PT_REQMT,pmi=PMI_REQMT,tiein=TIE_IN_NO,opttemp=OPTTEMP,subproject=SUB_PROJECT_NO,pcno=PC_No
                                                             ,ProId_id = id)
                
                messages.success(request, 'Data uploaded successfully.')
                return render(request, 'createproject.html')
            
            else:
                
                messages.error(request, 'Please upload a valid Excel file.')
                return render(request, 'createproject.html')
    else:
            return render (request,'upload_spoolgen.html')

    
    

def uploads_material(request,id):
    data = {
        "project_id": id, 
        "mir_list": MIRUpload.objects.filter(project__id=id)
    }
    print(data)
    return render(request,'uploads_material.html', data) 


def uploadmaterialslist(request, id):
    if request.method == 'POST':
        print(request.POST.copy())
        mir_upload = MIRUpload()
        mir_upload.location = request.POST.get('location')
        mir_upload.category = request.POST.get('category')
        mir_upload.mir_no = request.POST.get('mirNo')
        mir_upload.start_date = request.POST.get('startDate')
        mir_upload.part_id = request.POST.get('partId')
        mir_upload.mir_item = request.POST.get('item')
        mir_upload.material = request.POST.get('material')
        mir_upload.grade = request.POST.get('grade')
        mir_upload.weld_type = request.POST.get('WType')
        mir_upload.size_1 = request.POST.get('size1')
        mir_upload.sch_1 = request.POST.get('sch1')
        mir_upload.thk_1 = request.POST.get('thk1')
        mir_upload.size_2 = request.POST.get('size2')
        mir_upload.sch_2 = request.POST.get('sch2')
        mir_upload.thk_2 = request.POST.get('thk2')
        mir_upload.qty = request.POST.get('qty')
        mir_upload.length = request.POST.get('length')
        mir_upload.heat_no = request.POST.get('heatNo')
        mir_upload.certificate_no = request.POST.get('certiNo')
        mir_upload.mfr = request.POST.get('mfr')
        mir_upload.cal_qty = request.POST.get('callQty')
        mir_upload.project = project.objects.get(id=id)
        mir_upload.save()
        return redirect(reverse('uploads_material', args=(id,)))
    loc = weld_location.objects.all()
    weld_type_list = Weld.objects.all()
    mat_list = mat.objects.all()
    material_item_list = KVMaster.objects.filter(category="material_item").values('value')
    material_grade_list = KVMaster.objects.filter(category="material_grade", key=mat_list[0].materials).values('value')
    material_size_list = MaterialSizeThicknessMasterData.objects.all().values('size').distinct()
    material_schedule_list = MaterialSizeThicknessMasterData.objects.all().values('schedule').distinct()
    data = {
        "locations": loc,
        "weld_type_list": weld_type_list,
        "mat_list": mat_list,
        "material_item_list": material_item_list,
        "material_grade_list": material_grade_list,
        "material_size_list": material_size_list,
        "material_schedule_list": material_schedule_list
    }
    return render (request,'uploadmaterialslist.html', data)


def material_grade_fetch(request):
    material = request.GET.get('material')
    results = []
    if material and KVMaster.objects.filter(category="material_grade", key=material).exists():
        results = list(KVMaster.objects.filter(category="material_grade", key=material).values_list("value", flat=True))
    return JsonResponse({'results': results})


def material_schedule_fetch(request):
    size = request.GET.get('size')
    results = ""
    if size and MaterialSizeThicknessMasterData.objects.filter(size=size).exists():
        results = list(MaterialSizeThicknessMasterData.objects.filter(size=size).values_list("schedule", flat=True))
        thickness = MaterialSizeThicknessMasterData.objects.filter(size=size, schedule=results[0]).first().thickness
    return JsonResponse({'results': results, "thickness":thickness})


def material_thickness_fetch(request):
    size = request.GET.get('size')
    schedule = request.GET.get('schedule')
    results = ""
    if size and schedule and MaterialSizeThicknessMasterData.objects.filter(size=size, schedule=schedule).exists():
        results = MaterialSizeThicknessMasterData.objects.filter(size=size, schedule=schedule).first().thickness
    return JsonResponse({'results': results})


def part_number_detail_fetch(request):
    partno = request.GET.get('partno')
    results = {}
    if partno and MIRUpload.objects.filter(part_id=partno).exists():
        mir_upload = MIRUpload.objects.filter(part_id=partno).first()
        results = {
            'material': mir_upload.material,
            'heat_no': mir_upload.heat_no,
            'certificate_no': mir_upload.certificate_no
        }
    return JsonResponse({'results': results})

def uploads(request,id):
    projectId = id
    return render (request,'uploads.html',{'proId':projectId})


    